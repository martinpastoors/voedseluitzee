"""
Agrimatie Visserij Scraper
==========================
Haalt grafiekdata op van agrimatie.nl/visserij en exporteert
alle deelfiguren naar een Excel-bestand met één tabblad per pagina.

Gebruik:
    python scrape_agrimatie.py

Vereisten:
    pip install playwright openpyxl
    python -m playwright install chromium
"""

import json
import re
import time
from datetime import datetime
from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Pagina's om te scrapen ────────────────────────────────────────────────────
PAGES = [
    {
        "url": "https://agrimatie.nl/visserij/atotz/aanvoer-en-besomming-in-de-kottervisserij/",
        "tab": "Kottervisserij",
    },
    {
        "url": "https://agrimatie.nl/visserij/atotz/aanvoer-en-besomming-in-de-mosselcultuur/",
        "tab": "Mosselcultuur",
    },
    {
        "url": "https://agrimatie.nl/visserij/atotz/aanvoer-en-besomming-in-de-overige-kleine-zeevisserij/",
        "tab": "Overige kleine zeevisserij",
    },
    {
        "url": "https://agrimatie.nl/visserij/sectoren/grote-zeevisserij/",
        "tab": "Grote zeevisserij",
    },
]

OUTPUT_FILE = f"agrimatie_visserij_{datetime.today().strftime('%Y%m%d')}.xlsx"

# ── Stijlen ───────────────────────────────────────────────────────────────────
COLOR_HEADER_BG   = "1F4E79"   # donkerblauw
COLOR_HEADER_FG   = "FFFFFF"
COLOR_TITLE_BG    = "2E75B6"   # middelblauw
COLOR_ALT_ROW     = "D6E4F0"   # lichtblauw
COLOR_BORDER      = "8EAACC"

def thin_border():
    s = Side(style="thin", color=COLOR_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def header_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=True, color=COLOR_HEADER_FG, size=10)
    c.fill      = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = thin_border()
    return c

def title_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=True, color=COLOR_HEADER_FG, size=11)
    c.fill      = PatternFill("solid", fgColor=COLOR_TITLE_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    return c

def data_cell(ws, row, col, value, alt=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", size=10)
    c.alignment = Alignment(horizontal="right" if isinstance(value, (int, float)) else "left")
    c.border    = thin_border()
    if alt:
        c.fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)
    return c


# ── Datahaling via Playwright ─────────────────────────────────────────────────

def extract_chart_data_from_page(page):
    """
    Probeert op drie manieren grafiekdata te vinden:
    1. Onderschepte API-responses (JSON) opgeslagen in page.__api_data
    2. Highcharts / ApexCharts / Chart.js objecten in window
    3. SVG-tekst en tabel-elementen (fallback)
    """
    charts = []

    # ── Methode 1: Zoek naar Highcharts instanties ────────────────────────────
    try:
        hc_data = page.evaluate("""
        () => {
            const results = [];

            // Highcharts
            if (window.Highcharts && window.Highcharts.charts) {
                window.Highcharts.charts.forEach((chart, idx) => {
                    if (!chart) return;
                    const title   = chart.title?.textStr  || chart.options?.title?.text  || '';
                    const xTitle  = chart.xAxis?.[0]?.axisTitle?.textStr
                                 || chart.options?.xAxis?.title?.text || '';
                    const yTitle  = chart.yAxis?.[0]?.axisTitle?.textStr
                                 || chart.options?.yAxis?.title?.text || '';
                    const cats    = chart.xAxis?.[0]?.categories || [];
                    const series  = [];
                    chart.series.forEach(s => {
                        const pts = s.points
                            ? s.points.map(p => p.y ?? p.value ?? null)
                            : (s.options?.data || []);
                        series.push({ name: s.name || s.options?.name || '', data: pts });
                    });
                    results.push({ source: 'Highcharts', idx, title, xTitle, yTitle, categories: cats, series });
                });
            }
            return results;
        }
        """)
        if hc_data:
            charts.extend(hc_data)
    except Exception as e:
        print(f"    [Highcharts] {e}")

    # ── Methode 2: Zoek naar Vue/Nuxt component data in __vue__/__nuxt__ ──────
    if not charts:
        try:
            nuxt_data = page.evaluate("""
            () => {
                const results = [];
                // Zoek alle elementen met een __vue__ of __vueParentComponent property
                const allEls = document.querySelectorAll('*');
                const seen = new Set();
                for (const el of allEls) {
                    const vnode = el.__vue__ || el.__vueParentComponent?.proxy;
                    if (vnode && !seen.has(vnode)) {
                        seen.add(vnode);
                        try {
                            const data = JSON.stringify(vnode.$data || vnode._data || {});
                            if (data.length > 10 && (data.includes('series') || data.includes('categories') || data.includes('chartData'))) {
                                results.push({ source: 'Vue', snippet: data.slice(0, 2000) });
                            }
                        } catch {}
                    }
                }
                return results.slice(0, 10);
            }
            """)
            if nuxt_data:
                charts.extend(nuxt_data)
        except Exception as e:
            print(f"    [Vue] {e}")

    # ── Methode 3: Haal alle canvas/svg figuren op + bijschriften ─────────────
    try:
        dom_data = page.evaluate("""
        () => {
            const results = [];

            // SVG tekst per figuur
            const svgs = document.querySelectorAll('svg');
            svgs.forEach((svg, i) => {
                const texts = [...svg.querySelectorAll('text')].map(t => t.textContent.trim()).filter(Boolean);
                if (texts.length > 2) {
                    // Probeer tspan-elementen als series labels te gebruiken
                    results.push({ source: 'SVG', idx: i, texts });
                }
            });

            // HTML tabellen
            const tables = document.querySelectorAll('table');
            tables.forEach((tbl, i) => {
                const rows = [];
                tbl.querySelectorAll('tr').forEach(tr => {
                    const cells = [...tr.querySelectorAll('th,td')].map(c => c.textContent.trim());
                    rows.push(cells);
                });
                if (rows.length > 1) results.push({ source: 'Table', idx: i, rows });
            });

            return results;
        }
        """)
        if dom_data:
            charts.extend(dom_data)
    except Exception as e:
        print(f"    [DOM] {e}")

    return charts


def scrape_page(url, browser):
    """Open een pagina, wacht tot alle grafieken geladen zijn, extraheer data."""
    context = browser.new_context(
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/125.0.0.0 Safari/537.36"
        ),
        extra_http_headers={
            "Accept-Language": "nl-NL,nl;q=0.9",
            "Referer": "https://agrimatie.nl/visserij/",
        },
    )

    # Onderschep API-responses
    api_responses = []
    def on_response(resp):
        ct = resp.headers.get("content-type", "")
        if "json" in ct:
            try:
                body = resp.json()
                api_responses.append({"url": resp.url, "body": body})
            except:
                pass

    page = context.new_page()
    page.on("response", on_response)

    print(f"  Laden: {url}")
    page.goto(url, wait_until="networkidle", timeout=45000)

    # Wacht extra zodat lazy-loaded grafieken renderen
    page.wait_for_timeout(5000)

    # Scroll door de pagina zodat alle lazy-loaded elementen renderen
    page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
    page.wait_for_timeout(2000)
    page.evaluate("window.scrollTo(0, 0)")
    page.wait_for_timeout(1000)

    page_title = page.title()
    charts = extract_chart_data_from_page(page)

    context.close()
    return page_title, charts, api_responses


# ── Excel-export ──────────────────────────────────────────────────────────────

def write_highcharts_to_sheet(ws, charts, url):
    """Schrijft Highcharts-data naar een werkblad."""
    current_row = 1

    # Paginaheader
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    c = ws.cell(row=1, column=1, value=f"Bron: {url}")
    c.font = Font(name="Arial", italic=True, size=9, color="666666")
    ws.row_dimensions[1].height = 15
    current_row = 3

    hc_charts = [ch for ch in charts if ch.get("source") == "Highcharts"]
    table_charts = [ch for ch in charts if ch.get("source") == "Table"]

    if not hc_charts and not table_charts:
        ws.cell(row=current_row, column=1,
                value="⚠  Geen grafiekdata gevonden. Zie README voor alternatieve methode.")
        ws.cell(row=current_row, column=1).font = Font(name="Arial", color="CC0000", bold=True)
        return

    for chart in hc_charts:
        title = chart.get("title") or f"Figuur {chart.get('idx', '?') + 1}"
        cats  = chart.get("categories", [])
        series= chart.get("series", [])
        x_lbl = chart.get("xTitle", "Jaar/Categorie")
        y_lbl = chart.get("yTitle", "Waarde")

        if not series:
            continue

        # Figuur-titel
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=max(2, len(series) + 1))
        title_cell(ws, current_row, 1, title)
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # Kolomkoppen: [categorie/jaar] [serie1] [serie2] ...
        header_cell(ws, current_row, 1, x_lbl or "Jaar/Categorie")
        for si, s in enumerate(series):
            lbl = f"{s['name']}\n({y_lbl})" if y_lbl else s['name']
            header_cell(ws, current_row, si + 2, lbl)
        ws.row_dimensions[current_row].height = 28
        current_row += 1

        # Datarijen
        n_rows = max((len(s["data"]) for s in series), default=0)
        for ri in range(n_rows):
            alt = ri % 2 == 1
            cat_val = cats[ri] if ri < len(cats) else ri + 1
            data_cell(ws, current_row, 1, cat_val, alt)
            for si, s in enumerate(series):
                val = s["data"][ri] if ri < len(s["data"]) else None
                # Probeer numeriek te parsen
                if isinstance(val, str):
                    try:
                        val = float(val.replace(",", ".").replace(" ", ""))
                    except:
                        pass
                data_cell(ws, current_row, si + 2, val, alt)
            current_row += 1

        # Kolombreedtes
        ws.column_dimensions["A"].width = 18
        for si in range(len(series)):
            col_letter = get_column_letter(si + 2)
            ws.column_dimensions[col_letter].width = 18

        current_row += 2  # Ruimte tussen figuren

    # HTML-tabellen (fallback / aanvulling)
    for tbl in table_charts:
        rows = tbl.get("rows", [])
        if not rows:
            continue

        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=max(len(rows[0]), 2))
        title_cell(ws, current_row, 1, f"Tabel {tbl.get('idx', '') + 1}")
        current_row += 1

        for ri, row in enumerate(rows):
            alt = ri % 2 == 0
            for ci, val in enumerate(row):
                try:
                    num = float(val.replace(",", ".").replace(".", "", val.count(".") - 1))
                    cell_val = num
                except:
                    cell_val = val
                if ri == 0:
                    header_cell(ws, current_row, ci + 1, val)
                else:
                    data_cell(ws, current_row, ci + 1, cell_val, alt)
            current_row += 1

        current_row += 2


def write_fallback_info(ws, api_responses, url):
    """Als er geen Highcharts-data is maar wel API-responses, dump die als JSON."""
    ws.cell(row=3, column=1, value="API-responses onderschept (ruwe JSON):").font = Font(bold=True)
    r = 4
    for resp in api_responses[:5]:
        ws.cell(row=r, column=1, value=resp["url"])
        r += 1
        body_str = json.dumps(resp["body"], ensure_ascii=False, indent=2)[:5000]
        for line in body_str.split("\n"):
            ws.cell(row=r, column=1, value=line)
            r += 1
        r += 1


def build_readme_sheet(wb, pages_info):
    """Maak een README-tabblad met uitleg en overzicht."""
    ws = wb.create_sheet("README", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 70

    rows = [
        ("Agrimatie Visserij Data Export", None),
        (f"Gegenereerd op", datetime.today().strftime("%d-%m-%Y %H:%M")),
        ("", ""),
        ("Bron", "Wageningen Social & Economic Research / agrimatie.nl"),
        ("", ""),
        ("Tabbladen in dit bestand:", ""),
    ]
    for page in pages_info:
        rows.append((page["tab"], page["url"]))
    rows += [
        ("", ""),
        ("Methode", "Playwright headless browser + Highcharts API extractie"),
        ("", ""),
        ("Let op",
         "Als een tabblad '⚠ Geen grafiekdata gevonden' toont, \n"
         "probeer dan: (1) het script opnieuw te draaien, of \n"
         "(2) op de website zelf rechts-klikken op een grafiek → "
         "'Bekijk paginabron' en zoek naar 'csv' download-links."),
    ]

    for ri, (k, v) in enumerate(rows, 1):
        c1 = ws.cell(row=ri, column=1, value=k)
        c1.font = Font(name="Arial", bold=(ri == 1 or v is None or k.endswith(":")), size=10 if ri > 1 else 13)
        if ri == 1:
            c1.font = Font(name="Arial", bold=True, size=13, color=COLOR_TITLE_BG)
        if v is not None:
            c2 = ws.cell(row=ri, column=2, value=v)
            c2.font = Font(name="Arial", size=10)
            c2.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[ri].height = 16 if ri > 1 else 22


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  Agrimatie Visserij Scraper")
    print("=" * 60)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # verwijder standaard leeg blad

    pages_info = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)

        for page_cfg in PAGES:
            url   = page_cfg["url"]
            tab   = page_cfg["tab"]
            print(f"\n▶ {tab}")
            print(f"  URL: {url}")

            try:
                page_title, charts, api_responses = scrape_page(url, browser)
                print(f"  Paginatitel : {page_title or '(leeg)'}")
                print(f"  Highcharts  : {sum(1 for c in charts if c.get('source') == 'Highcharts')} figuren")
                print(f"  Tabellen    : {sum(1 for c in charts if c.get('source') == 'Table')} gevonden")
                print(f"  API-resp.   : {len(api_responses)} JSON-responses")

                ws = wb.create_sheet(tab[:31])
                ws.sheet_view.showGridLines = False

                hc = [c for c in charts if c.get("source") == "Highcharts"]
                if hc or [c for c in charts if c.get("source") == "Table"]:
                    write_highcharts_to_sheet(ws, charts, url)
                elif api_responses:
                    write_fallback_info(ws, api_responses, url)
                else:
                    ws.cell(row=3, column=1,
                            value="⚠  Geen data gevonden. Mogelijk blokkeert de site headless browsers.")
                    ws.cell(row=3, column=1).font = Font(name="Arial", color="CC0000", bold=True)
                    ws.cell(row=4, column=1,
                            value="Probeer het script opnieuw, of zie README voor alternatieve stappen.")

                pages_info.append({"tab": tab, "url": url, "charts": len(hc)})

            except Exception as e:
                print(f"  ❌ Fout: {e}")
                ws = wb.create_sheet(tab[:31])
                ws.cell(row=1, column=1, value=f"Fout bij laden: {e}")
                ws.cell(row=1, column=1).font = Font(color="CC0000")
                pages_info.append({"tab": tab, "url": url, "charts": 0})

        browser.close()

    build_readme_sheet(wb, pages_info)

    # Verplaats README naar eerste positie
    wb.move_sheet("README", offset=-len(wb.sheetnames) + 1)

    wb.save(OUTPUT_FILE)
    print(f"\n✅ Opgeslagen als: {OUTPUT_FILE}")
    total = sum(p["charts"] for p in pages_info)
    print(f"   Totaal figuren geëxporteerd: {total}")
    print("=" * 60)


if __name__ == "__main__":
    main()
